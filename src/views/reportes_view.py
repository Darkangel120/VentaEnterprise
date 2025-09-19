import flet as ft

def sort_table_by(app, criteria):
    """Sort the reportes table by the specified criteria"""
    try:
        # Set the sort criteria in the app
        app.sort_criteria = criteria

        # Update the view with sorted data
        app.main_column.controls[0] = app.build_main_content()
        app.page.update()

        # Show success message
        app.page.snack_bar = ft.SnackBar(
            content=ft.Text(f"Tabla ordenada por {criteria}"),
            bgcolor=ft.Colors.GREEN_600
        )
        app.page.snack_bar.open = True
        app.page.update()

    except Exception as e:
        print(f"Error sorting table: {e}")
        app.page.snack_bar = ft.SnackBar(
            content=ft.Text("Error al ordenar la tabla"),
            bgcolor=ft.Colors.RED_600
        )
        app.page.snack_bar.open = True
        app.page.update()

def export_reportes_excel(app, e):
    """Export reportes data to Excel file"""
    try:
        import os
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        estadisticas = app.venta_controller.get_estadisticas_reportes()

        if not estadisticas['productos_mas_vendidos']:
            app.page.snack_bar = ft.SnackBar(
                content=ft.Text("No hay datos para exportar"),
                bgcolor=ft.Colors.ORANGE_600
            )
            app.page.snack_bar.open = True
            app.page.update()
            return

        # Create exports directory if it doesn't exist outside the exe
        export_dir = os.path.join(os.path.expanduser("~"), "VentaEnterprise_exports")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{export_dir}/reportes_ventas_{timestamp}.xlsx"

        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Ventas"

        # Define headers
        headers = ['Producto', 'Ventas', 'Ingresos', 'Precio Promedio', 'Stock Restante', 'Porcentaje Ventas', 'Porcentaje Ingresos']
        ws.append(headers)

        # Apply styles to header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_ventas = sum(cantidad for _, cantidad, _ in estadisticas['productos_mas_vendidos'])
        total_ingresos = sum(ingresos for _, _, ingresos in estadisticas['productos_mas_vendidos'])

        # Write data rows
        for producto in estadisticas['productos_mas_vendidos']:
            nombre, cantidad, ingresos = producto

            # Get stock info
            producto_info = next((p for p in app.productos if p.nombre == nombre), None)
            stock_restante = producto_info.stock if producto_info else 0
            precio_promedio = ingresos / cantidad if cantidad > 0 else 0

            porcentaje_ventas = (cantidad / total_ventas * 100) if total_ventas > 0 else 0
            porcentaje_ingresos = (ingresos / total_ingresos * 100) if total_ingresos > 0 else 0

            ws.append([
                nombre,
                cantidad,
                ingresos,
                round(precio_promedio, 2),
                stock_restante,
                f"{porcentaje_ventas:.1f}%",
                f"{porcentaje_ingresos:.1f}%"
            ])

        # Add summary row
        ws.append([])
        ws.append([
            'TOTAL',
            total_ventas,
            total_ingresos,
            '',
            '',
            '100.0%',
            '100.0%'
        ])

        # Adjust column widths
        column_widths = [30, 10, 15, 18, 15, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Save workbook
        wb.save(filename)

        # Show success message with file path
        app.page.snack_bar = ft.SnackBar(
            content=ft.Text(f"Archivo Excel exportado: {filename}"),
            bgcolor=ft.Colors.GREEN_600,
            duration=5000
        )
        app.page.snack_bar.open = True
        app.page.update()

    except Exception as e:
        print(f"Error exporting Excel: {e}")
        app.page.snack_bar = ft.SnackBar(
            content=ft.Text("Error al exportar archivo Excel"),
            bgcolor=ft.Colors.RED_600
        )
        app.page.snack_bar.open = True
        app.page.update()

def get_color_with_opacity(color, opacity):
    """Convertir un color de Flet a uno con opacidad"""
    if color == ft.Colors.GREEN:
        return ft.Colors.GREEN_200
    elif color == ft.Colors.BLUE:
        return ft.Colors.BLUE_200
    elif color == ft.Colors.ORANGE:
        return ft.Colors.ORANGE_200
    elif color == ft.Colors.PURPLE:
        return ft.Colors.PURPLE_200
    elif color == ft.Colors.RED:
        return ft.Colors.RED_200
    elif color == ft.Colors.YELLOW:
        return ft.Colors.YELLOW_200
    elif color == ft.Colors.GREY:
        return ft.Colors.GREY_300
    elif color == ft.Colors.BLUE_GREY:
        return ft.Colors.BLUE_GREY_300
    else:
        return ft.Colors.BLUE_GREY_200

def create_metric_card(app, title, value, color, icon):
    # Crear color de fondo más oscuro
    bg_color = get_color_with_opacity(color, 0.7)

    # Colores adaptativos para texto
    text_color = ft.Colors.BLACK if not app.dark_mode else ft.Colors.WHITE

    # Iconos en colores base
    icon_color = color

    return ft.Container(
        content=ft.Column(
            controls=[
                ft.Container(
                    content=ft.Icon(icon, size=40, color=icon_color),
                    width=60,
                    height=60,
                    alignment=ft.alignment.center,
                    bgcolor=bg_color,
                    border_radius=15
                ),
                ft.Text(title, size=14, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER, color=text_color),
                ft.Text(value, size=24, weight=ft.FontWeight.BOLD, color=text_color, text_align=ft.TextAlign.CENTER),
            ],
            spacing=10,
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER
        ),
        width=320,
        height=160,
        padding=ft.padding.all(20),
        bgcolor=ft.Colors.WHITE if not app.dark_mode else ft.Colors.BLUE_GREY_900,
        border_radius=20,
        shadow=ft.BoxShadow(
            spread_radius=2,
            blur_radius=12,
            color=ft.Colors.BLACK12 if not app.dark_mode else ft.Colors.BLACK26
        )
    )

def build_reportes_view(app):
    # Obtener datos reales de la base de datos
    estadisticas = app.venta_controller.get_estadisticas_reportes()

    # Aplicar ordenamiento si está configurado
    if hasattr(app, 'sort_criteria') and app.sort_criteria:
        if app.sort_criteria == "ventas":
            estadisticas['productos_mas_vendidos'].sort(key=lambda x: x[1], reverse=True)
        elif app.sort_criteria == "ingresos":
            estadisticas['productos_mas_vendidos'].sort(key=lambda x: x[2], reverse=True)

    # Crear métricas dinámicas mejoradas
    metricas = []

    if estadisticas['productos_mas_vendidos']:
        producto_top = estadisticas['productos_mas_vendidos'][0]
        metricas.append(create_metric_card(
            app,
            "Producto Más Vendido",
            producto_top[0],  # nombre del producto
            ft.Colors.GREEN,
            ft.Icons.STAR
        ))
    else:
        metricas.append(create_metric_card(
            app,
            "Producto Más Vendido",
            "Sin datos",
            ft.Colors.GREY,
            ft.Icons.STAR
        ))

    metricas.append(create_metric_card(
        app,
        "Total de Ventas",
        str(estadisticas['total_ventas']),
        ft.Colors.BLUE,
        ft.Icons.SHOPPING_CART
    ))

    metricas.append(create_metric_card(
        app,
        "Promedio por Venta",
        f"${estadisticas['promedio_venta']:.2f}",
        ft.Colors.PURPLE,
        ft.Icons.ATTACH_MONEY
    ))

    # Calcular totales para porcentajes
    total_ventas = sum(cantidad for _, cantidad, _ in estadisticas['productos_mas_vendidos']) if estadisticas['productos_mas_vendidos'] else 0
    total_ingresos = sum(ingresos for _, _, ingresos in estadisticas['productos_mas_vendidos']) if estadisticas['productos_mas_vendidos'] else 0

    # Crear filas de la tabla de análisis mejorada
    table_rows = []
    if estadisticas['productos_mas_vendidos']:
        for producto in estadisticas['productos_mas_vendidos']:
            nombre, cantidad, ingresos = producto
            # Obtener stock restante del producto
            producto_info = next((p for p in app.productos if p.nombre == nombre), None)
            stock_restante = producto_info.stock if producto_info else 0
            precio_promedio = ingresos / cantidad if cantidad > 0 else 0

            # Calcular porcentajes
            porcentaje_ventas = (cantidad / total_ventas * 100) if total_ventas > 0 else 0
            porcentaje_ingresos = (ingresos / total_ingresos * 100) if total_ingresos > 0 else 0

            # Determinar tendencia (simulada - en producción usar datos históricos)
            tendencia = "↗️" if porcentaje_ventas > 15 else "➡️" if porcentaje_ventas > 5 else "↘️"
            tendencia_color = ft.Colors.GREEN if porcentaje_ventas > 15 else ft.Colors.ORANGE if porcentaje_ventas > 5 else ft.Colors.RED

            table_rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Row([
                    ft.Icon(ft.Icons.INVENTORY, size=16, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                    ft.Text(nombre, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)
                ], spacing=8)),
                ft.DataCell(ft.Column([
                    ft.Text(str(cantidad), weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                    ft.Text(f"{porcentaje_ventas:.1f}%", size=10, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)
                ], spacing=2, alignment=ft.MainAxisAlignment.CENTER)),
                ft.DataCell(ft.Column([
                    ft.Text(f"${ingresos:.2f}", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                    ft.Text(f"{porcentaje_ingresos:.1f}%", size=10, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)
                ], spacing=2, alignment=ft.MainAxisAlignment.CENTER)),
                ft.DataCell(ft.Text(f"${precio_promedio:.2f}", size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                ft.DataCell(ft.Container(
                    content=ft.Text(tendencia, size=16, color=ft.Colors.WHITE),
                    alignment=ft.alignment.center,
                    bgcolor=get_color_with_opacity(tendencia_color, 0.7),
                    padding=ft.padding.all(8),
                    border_radius=8
                )),
                ft.DataCell(ft.Container(
                    content=ft.Text(str(stock_restante), size=14, color=ft.Colors.WHITE),
                    alignment=ft.alignment.center,
                    bgcolor=ft.Colors.RED_400 if stock_restante < 10 else ft.Colors.GREEN_400 if stock_restante > 50 else ft.Colors.ORANGE_400,
                    padding=ft.padding.symmetric(horizontal=12, vertical=6),
                    border_radius=12
                )),
            ]))
    else:
        table_rows.append(ft.DataRow(cells=[
            ft.DataCell(ft.Text("No hay datos de ventas disponibles", col=6, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
            ft.DataCell(ft.Text("")),
            ft.DataCell(ft.Text("")),
            ft.DataCell(ft.Text("")),
            ft.DataCell(ft.Text("")),
            ft.DataCell(ft.Text("")),
        ]))

    # Agregar fila de resumen
    if estadisticas['productos_mas_vendidos']:
        table_rows.append(ft.DataRow(cells=[
            ft.DataCell(ft.Row([
                ft.Icon(ft.Icons.ANALYTICS, size=16, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                ft.Text("TOTAL", weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)
            ], spacing=8)),
            ft.DataCell(ft.Text(str(total_ventas), weight=ft.FontWeight.BOLD, size=16, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
            ft.DataCell(ft.Text(f"${total_ingresos:.2f}", weight=ft.FontWeight.BOLD, size=16, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
            ft.DataCell(ft.Text("—", size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
            ft.DataCell(ft.Text("📊", size=16, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
            ft.DataCell(ft.Text("—", size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
        ]))

    # Controles de filtro y ordenamiento
    filter_controls = ft.Container(
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.ElevatedButton(
                            "Ordenar por Ventas",
                            icon=ft.Icons.SORT,
                            on_click=lambda e: sort_table_by(app, "ventas"),
                            style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_600, color=ft.Colors.WHITE)
                        ),
                        ft.ElevatedButton(
                            "Ordenar por Ingresos",
                            icon=ft.Icons.SORT,
                            on_click=lambda e: sort_table_by(app, "ingresos"),
                            style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN_600, color=ft.Colors.WHITE)
                        ),
                        ft.ElevatedButton(
                            "Exportar Excel",
                            icon=ft.Icons.DOWNLOAD,
                            on_click=lambda e: export_reportes_excel(app, e),
                            style=ft.ButtonStyle(bgcolor=ft.Colors.ORANGE_600, color=ft.Colors.WHITE)
                        ),
                    ],
                    spacing=15,
                    alignment=ft.MainAxisAlignment.START
                )
            ],
            spacing=10
        ),
        padding=ft.padding.symmetric(vertical=10)
    )

    return ft.Column(
        controls=[
            # Header mejorado
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                ft.Text("📊 Reportes Detallados", size=32, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                                ft.Text("Análisis avanzado de ventas con métricas detalladas", size=16, color=ft.Colors.BLUE_GREY_600 if not app.dark_mode else ft.Colors.BLUE_GREY_400),
                            ],
                            spacing=5,
                            expand=True
                        ),
                        ft.Container(
                            content=ft.Column(
                                controls=[
                                    ft.Text(f"{len(estadisticas['productos_mas_vendidos']) if estadisticas['productos_mas_vendidos'] else 0}", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE),
                                    ft.Text("Productos Analizados", size=12, color=ft.Colors.BLUE_GREY_500 if not app.dark_mode else ft.Colors.BLUE_GREY_300),
                                ],
                                spacing=2,
                                alignment=ft.MainAxisAlignment.END
                            ),
                            padding=ft.padding.symmetric(horizontal=20, vertical=15),
                            bgcolor=ft.Colors.BLUE_50 if not app.dark_mode else ft.Colors.BLUE_900,
                            border_radius=15
                        )
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                padding=ft.padding.symmetric(vertical=20)
            ),

            ft.Divider(height=20),

            # Métricas Cards
            ft.Container(
                content=ft.Row(
                    controls=metricas,
                    spacing=20,
                    wrap=True
                ),
                padding=ft.padding.symmetric(vertical=10)
            ),

            ft.Divider(height=30),

            # Sección de tabla mejorada
            ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Row(
                            controls=[
                                ft.Text("📈 Análisis de Ventas Detallado", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK),
                                ft.Container(
                                    content=ft.Text("Datos en tiempo real", size=12, color=ft.Colors.BLUE_GREY_600 if not app.dark_mode else ft.Colors.BLUE_GREY_400),
                                    bgcolor=ft.Colors.BLUE_GREY_100 if not app.dark_mode else ft.Colors.BLUE_GREY_800,
                                    padding=ft.padding.symmetric(horizontal=12, vertical=6),
                                    border_radius=15
                                )
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                        ),

                        # Controles de filtro
                        filter_controls,

                        # Tabla mejorada
                        ft.Container(
                            content=ft.DataTable(
                                columns=[
                                    ft.DataColumn(ft.Text("Producto", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                    ft.DataColumn(ft.Text("Ventas", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                    ft.DataColumn(ft.Text("Ingresos", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                    ft.DataColumn(ft.Text("Precio Promedio", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                    ft.DataColumn(ft.Text("Tendencia", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                    ft.DataColumn(ft.Text("Stock", weight=ft.FontWeight.BOLD, size=14, color=ft.Colors.WHITE if app.dark_mode else ft.Colors.BLACK)),
                                ],
                                rows=table_rows,
                                expand=True,
                                border_radius=15,
                                heading_row_height=60,
                                data_row_min_height=55,
                                data_row_max_height=70,
                                heading_row_color=ft.Colors.BLUE_GREY_50 if not app.dark_mode else ft.Colors.BLUE_GREY_800,
                                data_row_color={
                                    "hovered": ft.Colors.BLUE_GREY_50 if not app.dark_mode else ft.Colors.BLUE_GREY_800
                                }
                            ),
                            expand=True,
                            padding=25,
                            bgcolor=ft.Colors.WHITE if not app.dark_mode else ft.Colors.BLUE_GREY_900,
                            border_radius=20,
                            shadow=ft.BoxShadow(
                                spread_radius=2,
                                blur_radius=15,
                                color=ft.Colors.BLACK12 if not app.dark_mode else ft.Colors.BLACK26
                            )
                        ),

                        # Leyenda
                        ft.Container(
                            content=ft.Row(
                                controls=[
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Text("↗️", size=14, color=ft.Colors.GREEN_50),
                                            ft.Text("Alta demanda", size=12, color=ft.Colors.GREEN_50)
                                        ], spacing=5),
                                        padding=ft.padding.symmetric(horizontal=10, vertical=5),
                                        bgcolor=ft.Colors.GREEN_500,
                                        border_radius=10
                                    ),
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Text("➡️", size=14, color=ft.Colors.ORANGE_50),
                                            ft.Text("Demanda media", size=12, color=ft.Colors.ORANGE_50)
                                        ], spacing=5),
                                        padding=ft.padding.symmetric(horizontal=10, vertical=5),
                                        bgcolor=ft.Colors.ORANGE_500,
                                        border_radius=10
                                    ),
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Text("↘️", size=14, color=ft.Colors.RED_50),
                                            ft.Text("Baja demanda", size=12, color=ft.Colors.RED_50)
                                        ], spacing=5),
                                        padding=ft.padding.symmetric(horizontal=10, vertical=5),
                                        bgcolor=ft.Colors.RED_500,
                                        border_radius=10
                                    ),
                                ],
                                spacing=15,
                                alignment=ft.MainAxisAlignment.CENTER
                            ),
                            padding=ft.padding.symmetric(vertical=15)
                        )
                    ],
                    spacing=20
                ),
                padding=ft.padding.symmetric(horizontal=20, vertical=10)
            )
        ],
        spacing=0,
        scroll=ft.ScrollMode.AUTO,
        expand=True
    )


